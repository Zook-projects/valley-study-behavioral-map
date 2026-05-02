#!/usr/bin/env python3
"""
build-placer.py — Convert the Placer.ai Visitor Origins sheet into JSON.

Reads:
  - ../Placer.ai - Zip Code Market Data.xlsx (relative to project root)
  - public/data/corridors.json — already-routed corridor graph from build-data.py

Writes (all under public/data/, additive — never touches LODES outputs):
  - placer-visitor-origins.json — one row per origin ZIP with both visits and
                                   visitors counts, scope, and corridorPath
  - placer-zips.json             — origin-side ZIP centroid + place metadata
  - placer-summary.json          — destination rollup (totals, top-N origin
                                   places, in-region vs out-of-region split,
                                   YoY headline)

The Placer Visitor Origins sheet contains two side-by-side tables:
  - Cols 1–14 ("Visits" table)    — total trips into Glenwood Springs (81601)
  - Cols 16–29 ("Visitors" table) — unique visitors (smaller numbers)

Both share dimensions: single destination 81601, single year 2025, nationwide
origin ZIPs across all 50 states + DC. Inner-joined on Origin Zipcode so each
output row carries both measures.

Local vs non-local scope:
  An origin is "local" when its haversine distance from Glenwood Springs
  (39.5505, -107.3248) is ≤ 75 miles, otherwise "non-local". Scope is
  independent of corridor routing — a 80xxx Denver ZIP routes through the
  East gateway and gets a corridorPath even though it is non-local; a
  geographically close UT ZIP without a 80/81 prefix is local but has no
  corridorPath. The Local-only filter on the frontend hides every non-local
  flow regardless of whether it is routable.

Corridor routing:
  Every origin that resolves to a graph node (anchor binding or 80xxx /
  81xxx prefix-based east/west gateway fallback) gets a corridorPath; the
  rest get an empty path and render as off-corridor dots when included.

Run via: python3 scripts/build-placer.py
"""

from __future__ import annotations

import heapq
import json
import math
import sys
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
OUT_DIR = PROJECT_ROOT / "public" / "data"
CORRIDORS_JSON = OUT_DIR / "corridors.json"
# The xlsx lives one directory above the project root (alongside the LEHD
# Zip Code Jobs Data file). Relative path so the build is portable.
XLSX_PATH = PROJECT_ROOT.parent / "Placer.ai - Zip Code Market Data.xlsx"

# ---------------------------------------------------------------------------
# Domain constants — kept local so this script doesn't import from anchors.py
# (parity with build-data.py's gateway logic is what matters, not the import).
# ---------------------------------------------------------------------------
DEST_ZIP = "81601"
DEST_PLACE = "Glenwood Springs"
DEST_NODE = "GWS"
DEST_LAT = 39.5505
DEST_LNG = -107.3248

# 75-mile haversine threshold separates "local" from "non-local" origins.
# Roughly captures the Roaring Fork Valley + Western Slope I-70 corridor
# (Aspen, Vail, Eagle, Grand Junction at the borderline) while excluding the
# Front Range (Denver ~150 mi, Colorado Springs ~190 mi).
LOCAL_RADIUS_MILES = 75.0
EARTH_RADIUS_MILES = 3958.7613

# Mirrors scripts/build-data.py — east of -107.3248 routes via I-70 East
# Gateway, west routes via I-70 West Gateway. Used as a fallback for 80xxx /
# 81xxx ZIPs that do not bind directly to an anchor or junction node.
GATEWAY_E_NODE = "GW_E"
GATEWAY_W_NODE = "GW_W"
GATEWAY_SPLIT_LNG = -107.3248

YOY_NULL_TOKENS = {"N/A", "Insignificant YOY change"}


# ---------------------------------------------------------------------------
# Corridor graph — read from already-built corridors.json (no OSRM, no geojson
# re-parse). This script's sole purpose is to attach corridorPath strings to
# each origin; the graph topology is invariant relative to build-data.py.
# ---------------------------------------------------------------------------
def load_graph() -> tuple[
    dict[str, dict],
    dict[str, list[tuple[str, str, float]]],
    dict[str, str],
]:
    if not CORRIDORS_JSON.exists():
        raise RuntimeError(
            f"corridors.json not found at {CORRIDORS_JSON} — "
            "run scripts/build-data.py first to generate it"
        )
    with CORRIDORS_JSON.open(encoding="utf-8") as fh:
        graph = json.load(fh)
    if graph.get("version") != 1:
        raise RuntimeError(
            f"unsupported corridors.json version: {graph.get('version')!r}"
        )

    nodes: dict[str, dict] = {n["id"]: n for n in graph["nodes"]}
    zip_to_node: dict[str, str] = {}
    for n in nodes.values():
        z = n.get("zip")
        if z:
            zip_to_node[str(z)] = n["id"]

    adjacency: dict[str, list[tuple[str, str, float]]] = {nid: [] for nid in nodes}
    for c in graph["corridors"]:
        cid = c["id"]
        a, b = c["from"], c["to"]
        length = float(c["lengthMeters"])
        adjacency[a].append((cid, b, length))
        adjacency[b].append((cid, a, length))

    return nodes, adjacency, zip_to_node


def haversine_miles(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    """Great-circle distance in statute miles between two (lat, lng) points."""
    rlat1 = math.radians(lat1)
    rlat2 = math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(rlat1) * math.cos(rlat2) * math.sin(dlng / 2) ** 2
    )
    return 2 * EARTH_RADIUS_MILES * math.asin(math.sqrt(a))


def shortest_corridor_path(
    adjacency: dict[str, list[tuple[str, str, float]]],
    start: str,
    end: str,
) -> list[str] | None:
    """Dijkstra by total length, tie-break by hop count, then path tuple.

    Mirrors scripts/build-data.py:shortest_corridor_path so visitor-route
    decisions match commuter-route decisions on the same graph.
    """
    if start == end:
        return []
    if start not in adjacency or end not in adjacency:
        return None
    pq: list[tuple[float, int, tuple[str, ...], str]] = [(0.0, 0, (), start)]
    best: dict[str, tuple[float, int, tuple[str, ...]]] = {start: (0.0, 0, ())}
    while pq:
        dist, hops, path, node = heapq.heappop(pq)
        if node == end:
            return list(path)
        prev = best.get(node)
        if prev is not None and (dist, hops, path) > prev:
            continue
        for cid, nbr, length in adjacency[node]:
            cand_dist = dist + length
            cand_hops = hops + 1
            cand_path = path + (cid,)
            existing = best.get(nbr)
            cand_key = (cand_dist, cand_hops, cand_path)
            if existing is None or cand_key < existing:
                best[nbr] = cand_key
                heapq.heappush(pq, (cand_dist, cand_hops, cand_path, nbr))
    return None


def resolve_origin_node(
    zip_code: str,
    state: str | None,
    lng: float | None,
    zip_to_node: dict[str, str],
) -> str | None:
    """Resolve an origin ZIP to a corridor-graph node.

    Direct binding (anchor / junction associated ZIP) → that node.
    80xxx or 81xxx prefix → gateway fallback by longitude (when lng known)
                            else by prefix (80 → east, 81 → west).
    Anything else (out-of-state, malformed) → None.
    """
    if zip_code in zip_to_node:
        return zip_to_node[zip_code]
    if not zip_code or not zip_code.isdigit() or len(zip_code) != 5:
        return None
    prefix = zip_code[:2]
    if prefix not in ("80", "81"):
        return None
    # Defense: only fall back to gateway routing when the state agrees the
    # origin is in Colorado. Placer's `Origin City` mapping occasionally
    # spills CO-prefix ZIPs into other states; in that case treat as
    # out-of-region rather than smearing them across the corridor graph.
    if state and state.upper() != "CO":
        return None
    if lng is not None:
        return GATEWAY_E_NODE if lng > GATEWAY_SPLIT_LNG else GATEWAY_W_NODE
    return GATEWAY_E_NODE if prefix == "80" else GATEWAY_W_NODE


# ---------------------------------------------------------------------------
# XLSX read — one pass over the Visitor Origins sheet. Columns are:
#
#  cols 1..14:  "Visits" half  — Destination_Key, Origin_Key, Year, Geo_Type,
#                                 Destination Zip, Origin Zip, Origin City,
#                                 State, lat, lng, % of Visits, Visits,
#                                 YoY Change in Visits, Data Notes
#  cols 16..29: "Visitors" half — same columns, Visitors-flavored
#
# Both halves are inner-joined on Origin Zipcode so each row in the output
# carries both measures.
# ---------------------------------------------------------------------------
def parse_yoy(raw: Any) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    s = str(raw).strip()
    if not s or s in YOY_NULL_TOKENS:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def coerce_zip(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return f"{int(raw):05d}"
    s = str(raw).strip()
    if not s or not s.isdigit():
        return None
    return s.zfill(5)


def coerce_float(raw: Any) -> float | None:
    if raw is None:
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def coerce_int(raw: Any) -> int | None:
    v = coerce_float(raw)
    return int(v) if v is not None else None


def read_visitor_origins() -> list[dict]:
    """Read both Visits and Visitors tables and inner-join on Origin Zipcode.

    Returns a list of dicts ordered by (originState, originZip) for
    deterministic output ordering.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required to read the Placer xlsx — "
            "install via `pip install openpyxl`"
        ) from e

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if "Visitor Origins" not in wb.sheetnames:
        raise RuntimeError(
            f"Visitor Origins sheet missing from {XLSX_PATH.name} — "
            f"available sheets: {wb.sheetnames}"
        )
    ws = wb["Visitor Origins"]

    # Headers live on row 4 (0-indexed row 3). Body rows start at row 5.
    visits_by_origin: dict[str, dict] = {}
    visitors_by_origin: dict[str, dict] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        cells = list(row)
        if len(cells) < 30:
            cells = cells + [None] * (30 - len(cells))

        # --- Visits half (cols 1..14, 1-indexed → list indices 1..14) -----
        v_dest = coerce_zip(cells[5])
        v_orig = coerce_zip(cells[6])
        if v_dest == DEST_ZIP and v_orig:
            visits_by_origin[v_orig] = {
                "originZip": v_orig,
                "originPlace": (cells[7] or "").strip() or None,
                "originState": (cells[8] or "").strip() or None,
                "lat": coerce_float(cells[9]),
                "lng": coerce_float(cells[10]),
                "share": coerce_float(cells[11]),
                "count": coerce_int(cells[12]),
                "yoy": parse_yoy(cells[13]),
            }

        # --- Visitors half (cols 16..29 → list indices 16..29) ------------
        u_dest = coerce_zip(cells[20])
        u_orig = coerce_zip(cells[21])
        if u_dest == DEST_ZIP and u_orig:
            visitors_by_origin[u_orig] = {
                "originZip": u_orig,
                "originPlace": (cells[22] or "").strip() or None,
                "originState": (cells[23] or "").strip() or None,
                "lat": coerce_float(cells[24]),
                "lng": coerce_float(cells[25]),
                "share": coerce_float(cells[26]),
                "count": coerce_int(cells[27]),
                "yoy": parse_yoy(cells[28]),
            }

    # Outer-union: keep every ZIP that appeared in EITHER table; the missing
    # measure is null on the merged row.
    all_zips = set(visits_by_origin) | set(visitors_by_origin)
    rows: list[dict] = []
    for z in all_zips:
        v = visits_by_origin.get(z)
        u = visitors_by_origin.get(z)
        # Prefer the visits-side metadata when both tables agree (the Placer
        # export is consistent across the two halves anyway, but defensive).
        meta = v or u
        if meta is None:
            continue
        rows.append({
            "originZip": z,
            "originPlace": meta.get("originPlace") or "",
            "originState": meta.get("originState") or "",
            "lat": meta.get("lat"),
            "lng": meta.get("lng"),
            "visits": v["count"] if v else None,
            "visitors": u["count"] if u else None,
            "visitsShare": v["share"] if v else None,
            "visitorsShare": u["share"] if u else None,
            "visitsYoY": v["yoy"] if v else None,
            "visitorsYoY": u["yoy"] if u else None,
        })

    rows.sort(key=lambda r: (r["originState"] or "", r["originZip"]))
    return rows


# ---------------------------------------------------------------------------
# Build outputs
# ---------------------------------------------------------------------------
def build_flow_rows(
    raw_rows: list[dict],
    adjacency: dict[str, list[tuple[str, str, float]]],
    zip_to_node: dict[str, str],
) -> tuple[list[dict], dict[str, Any]]:
    """Attach scope (distance-based) + corridorPath + boundNode to every row.

    `scope` is purely a distance test — local if origin within 75 miles of
    Glenwood Springs (haversine), non-local otherwise. `corridorPath` is
    independently computed for any origin that binds to a graph node, so a
    non-local Denver ZIP still gets a routable path through GW_E.
    """
    out: list[dict] = []
    stats = {
        "local": 0,
        "non_local": 0,
        "unknown_distance": 0,
        "self_loop": 0,
        "routed": 0,
        "no_path": 0,
        "off_graph": 0,
    }
    path_cache: dict[str, list[str] | None] = {}

    for r in raw_rows:
        ozip = r["originZip"]
        lat, lng = r["lat"], r["lng"]

        # ---- scope (distance from Glenwood Springs) ----------------------
        if ozip == DEST_ZIP:
            distance = 0.0
            scope = "local"
            stats["self_loop"] += 1
            stats["local"] += 1
        elif lat is None or lng is None:
            # No coordinates — can't compute distance. Treat as non-local
            # (conservative; keeps unverified origins out of the local
            # rollup).
            distance = None
            scope = "non-local"
            stats["unknown_distance"] += 1
            stats["non_local"] += 1
        else:
            distance = haversine_miles(lat, lng, DEST_LAT, DEST_LNG)
            if distance <= LOCAL_RADIUS_MILES:
                scope = "local"
                stats["local"] += 1
            else:
                scope = "non-local"
                stats["non_local"] += 1

        # ---- corridorPath (independent of scope) -------------------------
        if ozip == DEST_ZIP:
            corridor_path: list[str] = []
            bound_node: str | None = DEST_NODE
        else:
            bound_node = resolve_origin_node(
                ozip, r["originState"], lng, zip_to_node
            )
            if bound_node is None:
                corridor_path = []
                stats["off_graph"] += 1
            else:
                cached = path_cache.get(bound_node)
                if cached is None and bound_node not in path_cache:
                    cached = shortest_corridor_path(adjacency, bound_node, DEST_NODE)
                    path_cache[bound_node] = cached
                if cached is None:
                    corridor_path = []
                    bound_node = None
                    stats["no_path"] += 1
                else:
                    corridor_path = list(cached)
                    stats["routed"] += 1

        out.append({
            "originZip": ozip,
            "originPlace": r["originPlace"],
            "originState": r["originState"],
            "lat": lat,
            "lng": lng,
            "destZip": DEST_ZIP,
            "destPlace": DEST_PLACE,
            "year": 2025,
            "source": "Placer",
            "metrics": {
                "visits": r["visits"],
                "visitors": r["visitors"],
                "visitsShare": r["visitsShare"],
                "visitorsShare": r["visitorsShare"],
                "visitsYoY": r["visitsYoY"],
                "visitorsYoY": r["visitorsYoY"],
            },
            "scope": scope,
            "distanceMiles": round(distance, 2) if distance is not None else None,
            "corridorPath": corridor_path,
            "boundNode": bound_node,
        })
    return out, stats


def build_zip_meta(rows: list[dict]) -> list[dict]:
    """One entry per (originZip) — sorted, deterministic, no duplicates.

    Keeps zips.json (LODES) untouched; the visitor view consumes this file
    instead since the universe of origins is much larger and the LODES anchor
    concept does not apply to Placer data.
    """
    seen: dict[str, dict] = {}
    for r in rows:
        z = r["originZip"]
        if z in seen:
            continue
        seen[z] = {
            "zip": z,
            "place": r["originPlace"] or z,
            "state": r["originState"] or "",
            "lat": r["lat"],
            "lng": r["lng"],
            "scope": r["scope"],
            "boundNode": r["boundNode"],
        }
    # Always include the destination as a synthetic entry so the frontend
    # does not have to special-case it in lookups.
    if DEST_ZIP not in seen:
        seen[DEST_ZIP] = {
            "zip": DEST_ZIP,
            "place": DEST_PLACE,
            "state": "CO",
            "lat": DEST_LAT,
            "lng": DEST_LNG,
            "scope": "local",
            "boundNode": DEST_NODE,
        }
    return sorted(seen.values(), key=lambda x: (x["state"], x["zip"]))


def build_summary(rows: list[dict]) -> dict:
    """Destination-side rollup. Splits totals into in-region / out-of-region
    universes and surfaces a top-N origin-place rollup grouped by Origin City.
    """
    def total(field: str, scope: str | None = None) -> int:
        s = 0
        for r in rows:
            if scope and r["scope"] != scope:
                continue
            v = r["metrics"].get(field)
            if isinstance(v, (int, float)):
                s += v
        return int(s)

    def yoy_total(field: str, scope: str | None = None) -> int:
        s = 0
        for r in rows:
            if scope and r["scope"] != scope:
                continue
            v = r["metrics"].get(field)
            if isinstance(v, (int, float)):
                s += int(v)
        return s

    # Group by Origin City (place name) so multi-ZIP places (e.g., Grand
    # Junction's 81501/81504/81505/81506) collapse to one summary row.
    by_place: dict[str, dict] = {}
    for r in rows:
        if r["originZip"] == DEST_ZIP:
            continue
        place_key = (r["originPlace"] or r["originZip"]).strip()
        state = r["originState"] or ""
        # Disambiguate same-named places across different states.
        key = f"{place_key}, {state}" if state else place_key
        entry = by_place.setdefault(key, {
            "place": place_key,
            "state": state,
            "scope": r["scope"],  # promoted to 'local' below if any ZIP is local
            "zips": [],
            "visits": 0,
            "visitors": 0,
        })
        entry["zips"].append(r["originZip"])
        if r["scope"] == "local":
            entry["scope"] = "local"
        v = r["metrics"].get("visits") or 0
        u = r["metrics"].get("visitors") or 0
        entry["visits"] += int(v)
        entry["visitors"] += int(u)

    # Sort places by visits desc, take top 25 for headline use; keep all in a
    # secondary list so the dashboard can render an unbounded scroll.
    place_rows = sorted(
        by_place.values(),
        key=lambda x: (-x["visits"], -x["visitors"], x["place"]),
    )
    for p in place_rows:
        p["zips"].sort()

    return {
        "year": 2025,
        "destZip": DEST_ZIP,
        "destPlace": DEST_PLACE,
        "totals": {
            "visits": total("visits"),
            "visitors": total("visitors"),
            "visitsYoY": yoy_total("visitsYoY"),
            "visitorsYoY": yoy_total("visitorsYoY"),
        },
        "localRadiusMiles": LOCAL_RADIUS_MILES,
        "byScope": {
            "local": {
                "visits": total("visits", "local"),
                "visitors": total("visitors", "local"),
                "originCount": sum(1 for r in rows if r["scope"] == "local"),
            },
            "non-local": {
                "visits": total("visits", "non-local"),
                "visitors": total("visitors", "non-local"),
                "originCount": sum(1 for r in rows if r["scope"] == "non-local"),
            },
        },
        "topPlaces": place_rows[:25],
        "allPlaces": place_rows,
    }


# ---------------------------------------------------------------------------
# Emit — sorted keys + 2-space indent for byte-stable diffs
# ---------------------------------------------------------------------------
def write_json(path: Path, payload: Any) -> None:
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, sort_keys=True, ensure_ascii=False)
        fh.write("\n")


def main() -> int:
    if not XLSX_PATH.exists():
        print(
            f"  ! Placer workbook not found at {XLSX_PATH} — skipping placer build "
            "(this is non-fatal; LODES build is unaffected)",
            file=sys.stderr,
        )
        return 0

    print(f"→ loading corridor graph from {CORRIDORS_JSON.name}")
    nodes, adjacency, zip_to_node = load_graph()
    print(f"  → {len(nodes)} nodes, {len(zip_to_node)} ZIP→node bindings")

    print(f"→ reading {XLSX_PATH.name} · sheet 'Visitor Origins'")
    raw_rows = read_visitor_origins()
    print(f"  → {len(raw_rows)} unique origin ZIPs (visits ∪ visitors)")

    flow_rows, stats = build_flow_rows(raw_rows, adjacency, zip_to_node)
    print(
        f"  → scope: local={stats['local']} non-local={stats['non_local']} "
        f"unknown-distance={stats['unknown_distance']}"
    )
    print(
        f"  → routing: routed={stats['routed']} off-graph={stats['off_graph']} "
        f"no-path={stats['no_path']} self-loop={stats['self_loop']}"
    )

    zip_meta = build_zip_meta(flow_rows)
    summary = build_summary(flow_rows)

    print(f"→ writing 3 JSON files under {OUT_DIR}")
    write_json(OUT_DIR / "placer-visitor-origins.json", flow_rows)
    write_json(OUT_DIR / "placer-zips.json", zip_meta)
    write_json(OUT_DIR / "placer-summary.json", summary)
    print(
        f"  → flows: {len(flow_rows)} rows · "
        f"zips: {len(zip_meta)} entries · "
        f"top places: {len(summary['topPlaces'])}"
    )
    print(
        f"  → totals: visits={summary['totals']['visits']:,} "
        f"visitors={summary['totals']['visitors']:,}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
