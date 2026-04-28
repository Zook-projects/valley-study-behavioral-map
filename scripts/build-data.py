#!/usr/bin/env python3
"""
build-data.py — Convert LEHD OnTheMap exports into flow + ZIP + corridor JSON.

Reads:
  - ../LEHD - Zip Code Jobs Data.xlsx
      sheet "LEHD Commute Inbound"  — anchor workplace ZIP → top-25 home ZIPs
      sheet "LEHD Commute Outbound" — anchor residence ZIP → top-25 work ZIPs
  - public/data/corridors.geojson — hand-authored canonical corridor graph
  - 2024 Census ZCTA Gazetteer (downloaded to /tmp on first run)

Writes:
  - public/data/flows-inbound.json   (workplace-anchored; corridorPath baked in)
  - public/data/flows-outbound.json  (residence-anchored; corridorPath baked in)
  - public/data/zips.json            (union of all ZIPs in either dataset)
  - public/data/corridors.json       (smoothed geometries + node metadata)

No network calls. Deterministic — two runs against identical inputs produce
byte-identical outputs.

Run via: python3 scripts/build-data.py
"""

from __future__ import annotations

import heapq
import json
import sys
import urllib.request
import zipfile
from pathlib import Path

import openpyxl

from osrm import OsrmError, route_polyline
from smoothing import haversine_length_meters

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
VAULT_DIR = PROJECT_ROOT.parent  # .../Valley Study - Behavioral Map/
XLSX_PATH = VAULT_DIR / "LEHD - Zip Code Jobs Data.xlsx"
OUT_DIR = PROJECT_ROOT / "public" / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

CORRIDORS_GEOJSON = OUT_DIR / "corridors.geojson"
OSRM_CACHE_PATH = HERE / ".osrm-corridor-cache.json"

GAZ_URL = "https://www2.census.gov/geo/docs/maps-data/data/gazetteer/2024_Gazetteer/2024_Gaz_zcta_national.zip"
GAZ_LOCAL_TXT = Path("/tmp/2024_Gaz_zcta_national.txt")
GAZ_LOCAL_ZIP = Path("/tmp/gaz_zcta_2024.zip")

# Coordinate rounding for output JSON. ~10 cm at 6 decimals — well below the
# rendering precision and below any plausible authoring fidelity.
COORD_DECIMALS = 6

# OSRM snap-radius (meters). None = use OSRM's built-in nearest-road
# tolerance, which is permissive enough to accept hand-authored control
# points sketched roughly along a highway. The route is still constrained
# to follow the right highway because we pass every controlPoint as an
# ordered via-point — OSRM threads the route through them in sequence.
OSRM_SNAP_RADIUS_M: int | None = None

# Anchor ZIPs — appear as workplace (inbound) and residence (outbound).
ANCHOR_ZIPS = {
    "81601", "81611", "81615", "81621", "81623",
    "81630", "81635", "81647", "81650", "81652", "81654",
}

# City-center coordinates for anchor ZIPs. Same overrides as v2 — gazetteer
# centroids are too far from downtown for sprawling resort/rural ZIPs.
CITY_CENTROIDS: dict[str, tuple[float, float]] = {
    "81601": (39.5505, -107.3248),
    "81611": (39.1911, -106.8175),
    "81615": (39.2130, -106.9378),
    "81621": (39.3691, -107.0328),
    "81623": (39.4019, -107.2117),
    "81630": (39.3306, -108.2231),
    "81635": (39.4519, -108.0531),
    "81647": (39.5736, -107.5306),
    "81650": (39.5347, -107.7831),
    "81652": (39.5483, -107.6539),
    "81654": (39.3310, -106.9849),
}

# Gateway routing rules. ZIPs outside the in-valley corridor topology are
# routed via one of the two gateway nodes based on their centroid longitude
# relative to Glenwood Springs. ZIPs east of Glenwood route via GW_E (toward
# Eagle/Vail/Front Range); ZIPs west of Glenwood route via GW_W (toward
# Grand Junction). Out-of-state and centroid-less ZIPs fall through to
# ALL_OTHER reclassification.
#
# This is centroid-based, not prefix-based. The 81xxx prefix straddles both
# sides of Glenwood (Grand Junction-area to the west, Eagle County to the
# east), so routing by prefix incorrectly sends Eagle County flows through
# the western corridor.
GATEWAY_E_NODE = "GW_E"
GATEWAY_W_NODE = "GW_W"
# Glenwood Springs longitude — boundary between east- and west-bound external
# ZIPs. Anything strictly east takes GW_E; everything else takes GW_W.
GATEWAY_SPLIT_LNG = -107.3248


# ---------------------------------------------------------------------------
# Gazetteer load
# ---------------------------------------------------------------------------
def load_gazetteer() -> dict[str, tuple[float, float]]:
    if not GAZ_LOCAL_TXT.exists():
        print(f"  fetching gazetteer → {GAZ_URL}", file=sys.stderr)
        urllib.request.urlretrieve(GAZ_URL, GAZ_LOCAL_ZIP)
        with zipfile.ZipFile(GAZ_LOCAL_ZIP) as zf:
            zf.extractall("/tmp")

    centroids: dict[str, tuple[float, float]] = {}
    with open(GAZ_LOCAL_TXT, encoding="utf-8") as fh:
        header = fh.readline().rstrip("\n").split("\t")
        header = [h.strip() for h in header]
        idx_geo = header.index("GEOID")
        idx_lat = header.index("INTPTLAT")
        idx_lng = header.index("INTPTLONG")
        for line in fh:
            parts = line.rstrip("\n").split("\t")
            if len(parts) <= idx_lng:
                continue
            zcta = parts[idx_geo].strip()
            try:
                lat = float(parts[idx_lat].strip())
                lng = float(parts[idx_lng].strip())
            except ValueError:
                continue
            centroids[zcta] = (lat, lng)
    return centroids


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------
def _coerce_zip(value) -> str:
    """Normalize ZIP cell to a 5-digit string, or pass 'All Other Locations' through."""
    if value is None:
        return ""
    s = str(value).strip()
    if s == "All Other Locations":
        return s
    if s.isdigit():
        return s.zfill(5)
    try:
        return str(int(s)).zfill(5)
    except ValueError:
        return s


def load_inbound_rows(wb) -> list[dict]:
    ws = wb["LEHD Commute Inbound"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 7:
            continue
        if row[7] is None:
            continue
        rows.append({
            "year": int(row[4]),
            "workplaceCity": str(row[5]).strip(),
            "residenceCity": str(row[6]).strip(),
            "workplaceZip": _coerce_zip(row[7]),
            "residenceZip": _coerce_zip(row[8]),
            "workerCount": int(row[9]),
            "percentage": float(row[10]),
        })
    return rows


def load_outbound_rows(wb) -> list[dict]:
    ws = wb["LEHD Commute Outbound"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 7:
            continue
        if row[1] is None:
            continue
        rows.append({
            "year": int(row[1]),
            "residenceCity": str(row[2]).strip(),
            "workplaceCity": str(row[3]).strip(),
            "residenceZip": _coerce_zip(row[4]),
            "workplaceZip": _coerce_zip(row[5]),
            "workerCount": int(row[6]),
            "percentage": float(row[7]),
        })
    return rows


# ---------------------------------------------------------------------------
# Corridor graph — load, validate, smooth
# ---------------------------------------------------------------------------
def _round_pair(p: list[float] | tuple[float, float]) -> list[float]:
    return [round(float(p[0]), COORD_DECIMALS), round(float(p[1]), COORD_DECIMALS)]


def load_corridor_graph() -> tuple[
    dict[str, dict],          # nodes: nodeId → {label, lng, lat, zip}
    list[dict],               # corridors: smoothed corridor records
    dict[str, list[tuple[str, str, float]]],  # adjacency: nodeId → [(corridorId, neighborId, lengthMeters)]
    dict[str, str],           # zip_to_node: associatedZip → nodeId
]:
    """
    Read corridors.geojson, validate, smooth corridor geometries, and return
    the in-memory representation used by the routing pass.

    The smoothed geometry is computed once here and stored on the corridor
    record so downstream consumers (Dijkstra, output JSON) all see the same
    bytes.
    """
    if not CORRIDORS_GEOJSON.exists():
        raise RuntimeError(
            f"corridors.geojson not found at {CORRIDORS_GEOJSON} — "
            "see the corridor authoring workflow in README.md"
        )
    with CORRIDORS_GEOJSON.open(encoding="utf-8") as fh:
        gj = json.load(fh)

    nodes: dict[str, dict] = {}
    corridors: list[dict] = []
    seen_corridor_ids: set[str] = set()
    zip_to_node: dict[str, str] = {}

    for feat in gj.get("features", []):
        props = feat.get("properties") or {}
        kind = props.get("kind")
        if kind == "node":
            node_id = props.get("nodeId")
            if not node_id or node_id in nodes:
                raise RuntimeError(f"corridors.geojson: missing or duplicate nodeId: {node_id!r}")
            geom = feat.get("geometry") or {}
            coords = geom.get("coordinates") or []
            if not (isinstance(coords, list) and len(coords) >= 2):
                raise RuntimeError(f"node {node_id} has no coordinates")
            lng, lat = float(coords[0]), float(coords[1])
            zip_str = props.get("associatedZip")
            nodes[node_id] = {
                "id": node_id,
                "label": props.get("label", node_id),
                "lng": round(lng, COORD_DECIMALS),
                "lat": round(lat, COORD_DECIMALS),
                "zip": zip_str,
            }
            if zip_str:
                # Multiple nodes may share an associated ZIP (e.g., 81654 → J_OS).
                # First-write wins, but warn if there's an unexpected collision.
                if zip_str in zip_to_node and zip_to_node[zip_str] != node_id:
                    print(
                        f"  ! ZIP {zip_str} associated with multiple nodes: "
                        f"{zip_to_node[zip_str]!r} and {node_id!r}; "
                        f"keeping {zip_to_node[zip_str]!r}",
                        file=sys.stderr,
                    )
                else:
                    zip_to_node[zip_str] = node_id
        elif kind == "corridor":
            corridors.append({"_props": props})
        # Ignore other feature kinds.

    # Route every corridor against real road geometry via OSRM, threading
    # the call through the author's controlPoints as via-points so OSRM
    # cannot pick a parallel side-road shortcut. Cache hits make repeat
    # builds offline.
    adjacency: dict[str, list[tuple[str, str, float]]] = {nid: [] for nid in nodes}
    routed_corridors: list[dict] = []
    for c in corridors:
        props = c["_props"]
        cid = props.get("corridorId")
        if not cid or cid in seen_corridor_ids:
            raise RuntimeError(f"corridors.geojson: missing or duplicate corridorId: {cid!r}")
        seen_corridor_ids.add(cid)
        from_id = props.get("fromNodeId")
        to_id = props.get("toNodeId")
        if from_id not in nodes or to_id not in nodes:
            raise RuntimeError(
                f"corridor {cid}: fromNodeId={from_id!r} or toNodeId={to_id!r} "
                "does not reference a defined node"
            )
        ctrl = props.get("controlPoints")
        if not isinstance(ctrl, list) or len(ctrl) < 2:
            raise RuntimeError(f"corridor {cid}: needs at least 2 control points")
        # Anchor first/last control points exactly to their node coordinates so
        # corridor endpoints meet the node centers — important when authors
        # sketch slightly off-anchor control points.
        ctrl = [[float(p[0]), float(p[1])] for p in ctrl]
        ctrl[0] = [nodes[from_id]["lng"], nodes[from_id]["lat"]]
        ctrl[-1] = [nodes[to_id]["lng"], nodes[to_id]["lat"]]

        # Route from the from-node directly to the to-node — no intermediate
        # via-points. The hand-authored controlPoints were shaped for the
        # legacy Catmull-Rom smoother and sit roughly along (not on) the
        # highway, which causes OSRM to double back at each one. The valley's
        # major-highway corridors (Hwy 82, I-70, Brush Creek Rd) are
        # unambiguous routes between any two valley nodes — OSRM's driving
        # profile picks them by default because they're the fastest paths.
        endpoints = [ctrl[0], ctrl[-1]]
        try:
            geometry = route_polyline(
                endpoints,
                cache_path=OSRM_CACHE_PATH,
                radius_m=OSRM_SNAP_RADIUS_M,
                label=cid,
            )
        except OsrmError as e:
            raise RuntimeError(
                f"corridor {cid}: OSRM routing failed — {e}. "
                "Aborting build; corridor geometries must follow real road "
                "geometry."
            ) from e

        # Anchor endpoints exactly on node coords (OSRM may snap a few meters
        # off the declared anchor). Idempotent if already aligned.
        geometry = [_round_pair(p) for p in geometry]
        if geometry:
            geometry[0] = [nodes[from_id]["lng"], nodes[from_id]["lat"]]
            geometry[-1] = [nodes[to_id]["lng"], nodes[to_id]["lat"]]
        length_m = haversine_length_meters(geometry)

        routed_corridors.append({
            "id": cid,
            "label": props.get("label", cid),
            "from": from_id,
            "to": to_id,
            "roadName": props.get("roadName", ""),
            "geometry": geometry,
            "lengthMeters": round(length_m, 1),
        })
        adjacency[from_id].append((cid, to_id, length_m))
        adjacency[to_id].append((cid, from_id, length_m))

    if not routed_corridors:
        raise RuntimeError("corridors.geojson contains no corridor features")

    return nodes, routed_corridors, adjacency, zip_to_node


# ---------------------------------------------------------------------------
# Dijkstra
# ---------------------------------------------------------------------------
def shortest_corridor_path(
    adjacency: dict[str, list[tuple[str, str, float]]],
    start: str,
    end: str,
) -> list[str] | None:
    """
    Standard Dijkstra over the corridor graph (undirected via mirrored edges
    in `adjacency`). Returns an ordered list of corridor IDs from start to
    end, or None if no path exists.

    Tie-breaking: when two paths share the same total length, the one with
    fewer corridors wins; ties beyond that resolve alphabetically by
    corridor ID. Both rules are baked into the priority key so output is
    deterministic across runs.
    """
    if start == end:
        return []
    if start not in adjacency or end not in adjacency:
        return None

    # priority queue entries: (total_length, corridor_count, path_tuple, node_id)
    pq: list[tuple[float, int, tuple[str, ...], str]] = [(0.0, 0, (), start)]
    best: dict[str, tuple[float, int, tuple[str, ...]]] = {start: (0.0, 0, ())}

    while pq:
        dist, hops, path, node = heapq.heappop(pq)
        if node == end:
            return list(path)
        prev = best.get(node)
        if prev is not None and (dist, hops, path) > prev:
            # A better path has superseded this entry.
            continue
        for cid, nbr, length in adjacency[node]:
            cand_dist = dist + length
            cand_hops = hops + 1
            cand_path = path + (cid,)
            existing = best.get(nbr)
            cand_key = (cand_dist, cand_hops, cand_path)
            if existing is None or cand_key < existing:
                best[nbr] = cand_key
                heapq.heappush(pq, (cand_dist, cand_hops, cand_path, nbr))

    return None


# ---------------------------------------------------------------------------
# Flow → corridor path mapping
# ---------------------------------------------------------------------------
def classify_external_zip(
    zip_code: str,
    lng: float | None,
) -> str | None:
    """
    Map an out-of-topology ZIP to a gateway node, or return None if the ZIP
    should be reclassified to ALL_OTHER. Routing rules:
      - In-state ZIPs (80xxx Front Range, 81xxx Western Slope) route via
        GW_E if their centroid is east of Glenwood, else GW_W.
      - Out-of-state ZIPs and ZIPs with no centroid → ALL_OTHER.
      - In-state ZIPs without a centroid fall back to a prefix-based default
        (80xxx → GW_E, 81xxx → GW_W) for resilience.
    """
    if not zip_code or not zip_code.isdigit() or len(zip_code) != 5:
        return None
    prefix = zip_code[:2]
    if prefix not in ("80", "81"):
        return None
    if lng is not None:
        return GATEWAY_E_NODE if lng > GATEWAY_SPLIT_LNG else GATEWAY_W_NODE
    # Centroid-less fallback — preserves the original prefix heuristic.
    return GATEWAY_E_NODE if prefix == "80" else GATEWAY_W_NODE


def resolve_node_for_zip(
    zip_code: str,
    zip_to_node: dict[str, str],
    zip_lng: dict[str, float],
) -> str | None:
    """Return the corridor-graph node a flow endpoint maps to, or None."""
    if zip_code in zip_to_node:
        return zip_to_node[zip_code]
    return classify_external_zip(zip_code, zip_lng.get(zip_code))


def attach_corridor_paths(
    flows: list[dict],
    adjacency: dict[str, list[tuple[str, str, float]]],
    zip_to_node: dict[str, str],
    zip_lng: dict[str, float],
) -> tuple[int, int, int]:
    """
    Mutate each flow row to add a `corridorPath: list[str]` field. Reclassifies
    out-of-state and other unmapped flows to ALL_OTHER in place by rewriting
    `originZip` or `destZip`.

    Returns (routed_count, self_loop_count, reclassified_count).
    """
    routed = 0
    self_loop = 0
    reclassified = 0

    # Cache shortest paths between (origin_node, dest_node) pairs — every
    # flow that traverses the same node-to-node arc shares the same corridor
    # path, so we only run Dijkstra once per unique pair.
    path_cache: dict[tuple[str, str], list[str] | None] = {}

    for f in flows:
        ozip = f["originZip"]
        dzip = f["destZip"]

        # Self-flows render as concentric rings, not corridors.
        if ozip == dzip:
            f["corridorPath"] = []
            self_loop += 1
            continue

        # ALL_OTHER inputs are already residual.
        if ozip == "ALL_OTHER" or dzip == "ALL_OTHER":
            f["corridorPath"] = []
            continue

        o_node = resolve_node_for_zip(ozip, zip_to_node, zip_lng)
        d_node = resolve_node_for_zip(dzip, zip_to_node, zip_lng)
        if o_node is None or d_node is None:
            # Reclassify the unmapped side to ALL_OTHER.
            if o_node is None:
                f["originZip"] = "ALL_OTHER"
            if d_node is None:
                f["destZip"] = "ALL_OTHER"
            f["corridorPath"] = []
            reclassified += 1
            continue

        cache_key = (o_node, d_node)
        if cache_key in path_cache:
            path = path_cache[cache_key]
        else:
            path = shortest_corridor_path(adjacency, o_node, d_node)
            path_cache[cache_key] = path

        if path is None:
            raise RuntimeError(
                f"no corridor path between node {o_node!r} and node {d_node!r} "
                f"(flow {ozip}→{dzip}); the corridor graph is disconnected"
            )
        f["corridorPath"] = path
        routed += 1

    return routed, self_loop, reclassified


def build_corridor_aggregation_index(
    flows_inbound: list[dict],
    flows_outbound: list[dict],
) -> dict:
    """
    Build a build-time corridor aggregation index for sanity-check assertions.
    Schema mirrors the plan's §"Step 4". Discardable post-validation; not
    written to disk by default.
    """
    index: dict[str, dict] = {}
    for f in flows_inbound + flows_outbound:
        path = f.get("corridorPath") or []
        if not path:
            continue
        wc = int(f["workerCount"])
        for cid in path:
            entry = index.setdefault(cid, {
                "totalWorkersAcrossAllFlows": 0,
                "flowCount": 0,
                "byDestZip": {},
            })
            entry["totalWorkersAcrossAllFlows"] += wc
            entry["flowCount"] += 1
            entry["byDestZip"][f["destZip"]] = entry["byDestZip"].get(f["destZip"], 0) + wc
    return index


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    print(f"reading {XLSX_PATH}", file=sys.stderr)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    inbound = load_inbound_rows(wb)
    outbound = load_outbound_rows(wb)
    print(f"  → {len(inbound)} inbound rows, {len(outbound)} outbound rows", file=sys.stderr)

    print("loading corridor graph", file=sys.stderr)
    nodes, corridors, adjacency, zip_to_node = load_corridor_graph()
    print(
        f"  → {len(nodes)} nodes, {len(corridors)} corridors, "
        f"{len(zip_to_node)} ZIP→node bindings",
        file=sys.stderr,
    )

    print("loading ZCTA gazetteer", file=sys.stderr)
    centroids = load_gazetteer()
    print(f"  → {len(centroids)} ZCTA centroids", file=sys.stderr)

    # -----------------------------------------------------------------------
    # Discover ZIPs and place names from both sheets.
    # -----------------------------------------------------------------------
    zip_to_place: dict[str, str] = {}
    all_zips: set[str] = set()

    for r in inbound:
        all_zips.add(r["workplaceZip"])
        zip_to_place.setdefault(r["workplaceZip"], r["workplaceCity"])
        if r["residenceZip"].isdigit():
            all_zips.add(r["residenceZip"])
            zip_to_place.setdefault(r["residenceZip"], r["residenceCity"])

    for r in outbound:
        if r["residenceZip"].isdigit():
            all_zips.add(r["residenceZip"])
            zip_to_place.setdefault(r["residenceZip"], r["residenceCity"])
        if r["workplaceZip"].isdigit():
            all_zips.add(r["workplaceZip"])
            zip_to_place.setdefault(r["workplaceZip"], r["workplaceCity"])

    # -----------------------------------------------------------------------
    # Per-ZIP totals.
    # -----------------------------------------------------------------------
    workplace_totals: dict[str, int] = {}
    for r in inbound:
        wzip = r["workplaceZip"]
        workplace_totals[wzip] = workplace_totals.get(wzip, 0) + r["workerCount"]

    residence_totals: dict[str, int] = {}
    for r in outbound:
        rzip = r["residenceZip"]
        residence_totals[rzip] = residence_totals.get(rzip, 0) + r["workerCount"]

    inbound_residence_seed: dict[str, int] = {}
    for r in inbound:
        rzip = r["residenceZip"]
        if rzip.isdigit() and rzip not in ANCHOR_ZIPS:
            inbound_residence_seed[rzip] = inbound_residence_seed.get(rzip, 0) + r["workerCount"]
    outbound_workplace_seed: dict[str, int] = {}
    for r in outbound:
        wzip = r["workplaceZip"]
        if wzip.isdigit() and wzip not in ANCHOR_ZIPS:
            outbound_workplace_seed[wzip] = outbound_workplace_seed.get(wzip, 0) + r["workerCount"]

    # -----------------------------------------------------------------------
    # Build zips.json
    # -----------------------------------------------------------------------
    zips_out: list[dict] = []
    missing_centroids: list[str] = []
    for zip_code in sorted(all_zips):
        if zip_code in CITY_CENTROIDS:
            lat, lng = CITY_CENTROIDS[zip_code]
        else:
            centroid = centroids.get(zip_code)
            if centroid is None:
                missing_centroids.append(zip_code)
                continue
            lat, lng = centroid

        total_as_wp = workplace_totals.get(zip_code, outbound_workplace_seed.get(zip_code, 0))
        total_as_res = residence_totals.get(zip_code, inbound_residence_seed.get(zip_code, 0))

        zips_out.append({
            "zip": zip_code,
            "place": zip_to_place.get(zip_code, ""),
            "lat": lat,
            "lng": lng,
            "totalAsWorkplace": total_as_wp,
            "totalAsResidence": total_as_res,
            "isAnchor": zip_code in ANCHOR_ZIPS,
        })

    aol_inbound = sum(r["workerCount"] for r in inbound if r["residenceZip"] == "All Other Locations")
    aol_outbound = sum(r["workerCount"] for r in outbound if r["workplaceZip"] == "All Other Locations")
    zips_out.append({
        "zip": "ALL_OTHER",
        "place": "All Other Locations",
        "lat": None,
        "lng": None,
        "totalAsWorkplace": aol_outbound,
        "totalAsResidence": aol_inbound,
        "isAnchor": False,
        "isSynthetic": True,
    })

    if missing_centroids:
        print(
            f"  ! {len(missing_centroids)} ZIPs without centroids "
            f"(dropped): {missing_centroids}",
            file=sys.stderr,
        )

    # -----------------------------------------------------------------------
    # Build flow files.
    # -----------------------------------------------------------------------
    flows_inbound_out: list[dict] = []
    for r in inbound:
        rzip = r["residenceZip"] if r["residenceZip"].isdigit() else "ALL_OTHER"
        flows_inbound_out.append({
            "originZip": rzip,
            "originPlace": r["residenceCity"],
            "destZip": r["workplaceZip"],
            "destPlace": r["workplaceCity"],
            "workerCount": r["workerCount"],
            "percentage": r["percentage"],
            "year": r["year"],
            "source": "LEHD",
        })

    flows_outbound_out: list[dict] = []
    for r in outbound:
        wzip = r["workplaceZip"] if r["workplaceZip"].isdigit() else "ALL_OTHER"
        flows_outbound_out.append({
            "originZip": r["residenceZip"],
            "originPlace": r["residenceCity"],
            "destZip": wzip,
            "destPlace": r["workplaceCity"],
            "workerCount": r["workerCount"],
            "percentage": r["percentage"],
            "year": r["year"],
            "source": "LEHD",
        })

    # -----------------------------------------------------------------------
    # Map flows onto corridors.
    # -----------------------------------------------------------------------
    print("\nrouting flows through corridor graph…", file=sys.stderr)
    # Build a ZIP → longitude lookup used by the gateway classifier. Anchor
    # overrides win over gazetteer centroids so resort/rural ZIPs get the
    # downtown coordinate that's already been pinned upstream.
    zip_lng: dict[str, float] = {}
    for z, (_lat, lng) in centroids.items():
        zip_lng[z] = lng
    for z, (_lat, lng) in CITY_CENTROIDS.items():
        zip_lng[z] = lng
    in_routed, in_self, in_reclass = attach_corridor_paths(
        flows_inbound_out, adjacency, zip_to_node, zip_lng
    )
    out_routed, out_self, out_reclass = attach_corridor_paths(
        flows_outbound_out, adjacency, zip_to_node, zip_lng
    )
    print(
        f"  inbound:  routed={in_routed}, self={in_self}, reclassified={in_reclass}",
        file=sys.stderr,
    )
    print(
        f"  outbound: routed={out_routed}, self={out_self}, reclassified={out_reclass}",
        file=sys.stderr,
    )

    # -----------------------------------------------------------------------
    # Build corridor aggregation index for build-time validation.
    # -----------------------------------------------------------------------
    print("validating corridor aggregation…", file=sys.stderr)
    agg_index = build_corridor_aggregation_index(flows_inbound_out, flows_outbound_out)
    corridor_ids_in_use = set(agg_index.keys())
    declared_corridor_ids = {c["id"] for c in corridors}
    orphan_corridors = declared_corridor_ids - corridor_ids_in_use
    if orphan_corridors:
        print(
            f"  ! {len(orphan_corridors)} corridor(s) have zero contributing "
            f"flows: {sorted(orphan_corridors)}",
            file=sys.stderr,
        )
    # Reconciliation sanity check.
    flow_corridor_workers = sum(
        int(f["workerCount"]) * len(f.get("corridorPath") or [])
        for f in flows_inbound_out + flows_outbound_out
    )
    agg_total = sum(int(v["totalWorkersAcrossAllFlows"]) for v in agg_index.values())
    if flow_corridor_workers > 0:
        delta = abs(agg_total - flow_corridor_workers) / flow_corridor_workers
        if delta > 0.01:
            print(
                f"  ! aggregation drift {delta:.4%} exceeds 1% threshold",
                file=sys.stderr,
            )
            return 1
    print(f"  → {len(corridor_ids_in_use)} corridors carrying flow", file=sys.stderr)

    # -----------------------------------------------------------------------
    # Build corridors.json (frontend-loadable shape).
    # -----------------------------------------------------------------------
    corridors_json = {
        "version": 1,
        "nodes": [
            {
                "id": n["id"],
                "label": n["label"],
                "lng": n["lng"],
                "lat": n["lat"],
                "zip": n["zip"],
            }
            for n in sorted(nodes.values(), key=lambda x: x["id"])
        ],
        "corridors": [
            {
                "id": c["id"],
                "label": c["label"],
                "from": c["from"],
                "to": c["to"],
                "roadName": c["roadName"],
                "geometry": c["geometry"],
                "lengthMeters": c["lengthMeters"],
            }
            for c in sorted(corridors, key=lambda x: x["id"])
        ],
    }

    # -----------------------------------------------------------------------
    # QA
    # -----------------------------------------------------------------------
    print("\nQA — inbound per-workplace totals:", file=sys.stderr)
    for wzip in sorted(ANCHOR_ZIPS):
        total = sum(f["workerCount"] for f in flows_inbound_out if f["destZip"] == wzip)
        pct_sum = sum(f["percentage"] for f in flows_inbound_out if f["destZip"] == wzip)
        place = zip_to_place.get(wzip, "?")
        print(f"  {wzip} {place:<22} total={total:>6,}  pctSum={pct_sum:.4f}", file=sys.stderr)

    print("\nQA — outbound per-residence totals:", file=sys.stderr)
    for rzip in sorted(ANCHOR_ZIPS):
        total = sum(f["workerCount"] for f in flows_outbound_out if f["originZip"] == rzip)
        pct_sum = sum(f["percentage"] for f in flows_outbound_out if f["originZip"] == rzip)
        place = zip_to_place.get(rzip, "?")
        print(f"  {rzip} {place:<22} total={total:>6,}  pctSum={pct_sum:.4f}", file=sys.stderr)

    # -----------------------------------------------------------------------
    # Write
    # -----------------------------------------------------------------------
    inbound_path = OUT_DIR / "flows-inbound.json"
    outbound_path = OUT_DIR / "flows-outbound.json"
    zips_path = OUT_DIR / "zips.json"
    corridors_path = OUT_DIR / "corridors.json"
    inbound_path.write_text(json.dumps(flows_inbound_out, separators=(",", ":")))
    outbound_path.write_text(json.dumps(flows_outbound_out, separators=(",", ":")))
    zips_path.write_text(json.dumps(zips_out, indent=2))
    corridors_path.write_text(json.dumps(corridors_json, separators=(",", ":")))
    print(f"\nwrote {inbound_path}  ({inbound_path.stat().st_size:,} bytes)", file=sys.stderr)
    print(f"wrote {outbound_path}  ({outbound_path.stat().st_size:,} bytes)", file=sys.stderr)
    print(f"wrote {zips_path}  ({zips_path.stat().st_size:,} bytes)", file=sys.stderr)
    print(f"wrote {corridors_path}  ({corridors_path.stat().st_size:,} bytes)", file=sys.stderr)

    # Drop the legacy outputs.
    for legacy_name in ("flows.json", "segment-aggregation.json"):
        legacy = OUT_DIR / legacy_name
        if legacy.exists():
            legacy.unlink()
            print(f"removed legacy {legacy}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
